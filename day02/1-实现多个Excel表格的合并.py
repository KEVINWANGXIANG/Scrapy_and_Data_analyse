

from openpyxl.reader.excel import load_workbook
from collections import OrderedDict
from pyexcel_xls import save_data
from pyexcel_xls import get_data



def readXlsxFile(path):
    # dic={}
    file=load_workbook(filename=path)
    # print(file.get_sheet_names())
    sheets=file.get_sheet_names()
    allInfo=[]
    for sheetName in sheets:
        sheet=file.get_sheet_by_name(sheetName)
        sheetInfo=[]
        for lineNum in range(1, sheet.max_row+1):
            lineList=[]
            for columnNum in range(1, sheet.max_column+1):
                value=sheet.cell(row=lineNum, column=columnNum).value
                lineList.append(value)
            sheetInfo.append(value)
        allInfo.append(sheetInfo)
    return allInfo
        # dic[sheetName]=sheetInfo
    # return dic


def makeExcelFile(path,data):
    dic=OrderedDict()
    for sheetName,sheetValue in data.items():
        d={}
        d[sheetName]=sheetValue
        dic.update(d)
    save_data(path,dic)

# def makeExcelFile(path,data):
#     dic=OrderedDict()
#     for sheetName,sheetValue in data:
#         d={}
#         d[sheetName]=sheetValue
#         dic.update(d)
#     save_data(path,dic)


path=r"C:\Users\Administrator\Desktop\sunck.xlsx"
data=readXlsxFile(path)
print(data)
toPath=r"C:\Users\Administrator\Desktop\sunck1.xls"
makeExcelFile(toPath, {"表1": data})
print("******")
